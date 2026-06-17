import io
import logging
import os
from pathlib import Path

logger = logging.getLogger(__name__)

IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".gif", ".webp"}
PDF_EXTENSIONS = {".pdf"}
DOCX_EXTENSIONS = {".docx"}
DOC_EXTENSIONS = {".doc"}
SPREADSHEET_EXTENSIONS = {".xls", ".xlsx"}


def extract_text_from_file(*, file_path: str | Path, extension: str, content_type: str = "") -> str:
    ext = extension.lower()
    if ext in PDF_EXTENSIONS:
        return _extract_pdf_text(file_path)
    if ext in DOCX_EXTENSIONS:
        return _extract_docx_text(file_path)
    if ext in DOC_EXTENSIONS:
        return _extract_doc_text(file_path)
    if ext in SPREADSHEET_EXTENSIONS:
        return _extract_spreadsheet_text(file_path)
    if ext in IMAGE_EXTENSIONS:
        return _extract_image_text(file_path)
    raise ValueError(f"Formato non supportato per OCR: {ext}")


def extract_text_from_storage_file(file_field) -> str:
    extension = os.path.splitext(file_field.name)[1].lower()
    with file_field.open("rb") as uploaded:
        content = uploaded.read()
    buffer = io.BytesIO(content)
    buffer.name = file_field.name
    return extract_text_from_bytes(
        data=content,
        filename=file_field.name,
        extension=extension,
    )


def extract_text_from_bytes(*, data: bytes, filename: str, extension: str) -> str:
    ext = extension.lower()
    if ext in PDF_EXTENSIONS:
        return _extract_pdf_bytes(data)
    if ext in DOCX_EXTENSIONS:
        return _extract_docx_bytes(data)
    if ext in DOC_EXTENSIONS:
        return _extract_doc_bytes(data)
    if ext in SPREADSHEET_EXTENSIONS:
        return _extract_spreadsheet_bytes(data, ext)
    if ext in IMAGE_EXTENSIONS:
        return _extract_image_bytes(data)
    raise ValueError(f"Formato non supportato per OCR: {ext}")


def _extract_pdf_text(file_path: str | Path) -> str:
    with open(file_path, "rb") as handle:
        return _extract_pdf_bytes(handle.read())


def _extract_pdf_bytes(data: bytes) -> str:
    try:
        text = _extract_pdf_with_pypdf(data)
    except ValueError:
        raise
    except Exception as exc:
        logger.warning("Lettura PDF nativa fallita: %s", exc)
        text = ""
    if text.strip():
        return text
    return _extract_pdf_with_ocr(data)


def _extract_pdf_with_pypdf(data: bytes) -> str:
    try:
        from pypdf import PdfReader
    except ImportError:
        logger.warning("pypdf non disponibile")
        return ""

    try:
        reader = PdfReader(io.BytesIO(data))
    except Exception as exc:
        raise ValueError(
            "File PDF non valido o corrotto: impossibile leggere il contenuto."
        ) from exc

    pages = []
    for page in reader.pages:
        page_text = page.extract_text() or ""
        if page_text.strip():
            pages.append(page_text.strip())
    return "\n\n".join(pages)


def _extract_pdf_with_ocr(data: bytes) -> str:
    try:
        from pdf2image import convert_from_bytes
        import pytesseract
    except ImportError:
        logger.warning("pdf2image/pytesseract non disponibili per OCR PDF")
        return ""

    try:
        images = convert_from_bytes(data, dpi=200)
    except Exception:
        logger.exception("Conversione PDF in immagini fallita")
        return ""

    pages = []
    for image in images:
        pages.append(pytesseract.image_to_string(image, lang="ita+eng"))
    return "\n\n".join(page.strip() for page in pages if page.strip())


def _extract_docx_text(file_path: str | Path) -> str:
    with open(file_path, "rb") as handle:
        return _extract_docx_bytes(handle.read())


def _extract_docx_bytes(data: bytes) -> str:
    try:
        from docx import Document as DocxDocument
    except ImportError:
        logger.warning("python-docx non disponibile")
        return ""

    document = DocxDocument(io.BytesIO(data))
    paragraphs = [paragraph.text.strip() for paragraph in document.paragraphs if paragraph.text.strip()]
    return "\n".join(paragraphs)


def _extract_doc_text(file_path: str | Path) -> str:
    try:
        import subprocess
        import tempfile

        with tempfile.TemporaryDirectory() as tmpdir:
            input_path = Path(tmpdir) / "input.doc"
            input_path.write_bytes(Path(file_path).read_bytes())
            output_path = Path(tmpdir) / "input.txt"
            result = subprocess.run(
                ["antiword", str(input_path)],
                capture_output=True,
                text=True,
                check=False,
            )
            if result.returncode == 0 and result.stdout.strip():
                return result.stdout.strip()
    except Exception:
        logger.exception("Estrazione .doc con antiword fallita")

    logger.warning("Formato .doc non supportato senza antiword installato")
    return ""


def _extract_doc_bytes(data: bytes) -> str:
    try:
        import subprocess
        import tempfile

        with tempfile.TemporaryDirectory() as tmpdir:
            input_path = Path(tmpdir) / "input.doc"
            input_path.write_bytes(data)
            result = subprocess.run(
                ["antiword", str(input_path)],
                capture_output=True,
                text=True,
                check=False,
            )
            if result.returncode == 0 and result.stdout.strip():
                return result.stdout.strip()
    except Exception:
        logger.exception("Estrazione .doc con antiword fallita")
    return ""


def _extract_spreadsheet_text(file_path: str | Path, extension: str) -> str:
    with open(file_path, "rb") as handle:
        return _extract_spreadsheet_bytes(handle.read(), extension)


def _extract_spreadsheet_bytes(data: bytes, extension: str) -> str:
    try:
        from openpyxl import load_workbook
    except ImportError:
        logger.warning("openpyxl non disponibile")
        return ""

    workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    rows = []
    for sheet in workbook.worksheets:
        rows.append(f"[Foglio: {sheet.title}]")
        for row in sheet.iter_rows(values_only=True):
            values = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
            if values:
                rows.append(" | ".join(values))
    return "\n".join(rows)


def _extract_image_text(file_path: str | Path) -> str:
    with open(file_path, "rb") as handle:
        return _extract_image_bytes(handle.read())


def _extract_image_bytes(data: bytes) -> str:
    try:
        import pytesseract
        from PIL import Image
    except ImportError:
        logger.warning("pytesseract/Pillow non disponibili")
        return ""

    image = Image.open(io.BytesIO(data))
    return pytesseract.image_to_string(image, lang="ita+eng").strip()
