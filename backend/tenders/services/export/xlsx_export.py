from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .constants import (
    MATRICE_REQUISITI,
    OFFERTA_ECONOMICA,
    RELAZIONE_TECNICA,
    REPORT_PARTECIPABILITA,
    SCHEDA_GARA,
)
from .collector import TenderExportContext

HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")
HEADER_FONT = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
BODY_FONT = Font(name="Calibri", size=10)
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="0F172A")


def _autosize_columns(sheet, max_width: int = 48) -> None:
    for column_cells in sheet.columns:
        letter = get_column_letter(column_cells[0].column)
        length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[letter].width = min(max(length + 2, 12), max_width)


def _style_header_row(sheet, row: int = 1) -> None:
    for cell in sheet[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="top", wrap_text=True)


def _save_workbook(workbook: Workbook) -> bytes:
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def export_scheda_xlsx(ctx: TenderExportContext) -> bytes:
    scheda = ctx.scheda
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Scheda gara"
    sheet["A1"] = f"Scheda gara – CIG {scheda['cig']}"
    sheet["A1"].font = TITLE_FONT
    sheet.merge_cells("A1:B1")

    rows = [
        ("Campo", "Valore"),
        ("CIG", scheda["cig"]),
        ("CPV", scheda["cpv"]),
        ("Oggetto", scheda.get("oggetto", "")),
        ("Importo", scheda["importo"]),
        ("Scadenza", scheda["scadenza"]),
        ("Stato", scheda["stato"]),
        ("Fase", scheda["fase"]),
        ("Fonte", scheda["source"]),
        ("Priorità", scheda["priorita"]),
        ("Punteggio priorità", scheda.get("priority_score", "")),
        ("Organizzazione", scheda.get("organization", "")),
        ("Estrazione AI", "Sì" if scheda.get("ai_extracted") else "No"),
        ("Generato il", ctx.exported_at),
    ]
    for row_index, row in enumerate(rows, start=3):
        sheet.cell(row=row_index, column=1, value=row[0]).font = BODY_FONT
        sheet.cell(row=row_index, column=2, value=row[1]).font = BODY_FONT
    _style_header_row(sheet, 3)
    _autosize_columns(sheet)
    return _save_workbook(workbook)


def export_matrix_xlsx(ctx: TenderExportContext) -> bytes:
    matrix = ctx.matrix
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Matrice requisiti"
    sheet["A1"] = f"Matrice requisiti – CIG {matrix.tender_cig}"
    sheet["A1"].font = TITLE_FONT

    headers = ["ID", "Descrizione", "Categoria", "Tipo", "Soglia minima"] + [
        company["name"] for company in matrix.companies
    ]
    for col_index, header in enumerate(headers, start=1):
        sheet.cell(row=3, column=col_index, value=header)
    _style_header_row(sheet, 3)

    for row_index, requirement in enumerate(matrix.requirements, start=4):
        sheet.cell(row=row_index, column=1, value=requirement.requirement_id).font = BODY_FONT
        sheet.cell(row=row_index, column=2, value=requirement.descrizione).font = BODY_FONT
        sheet.cell(row=row_index, column=3, value=requirement.categoria_label).font = BODY_FONT
        sheet.cell(row=row_index, column=4, value=requirement.tipo_label).font = BODY_FONT
        sheet.cell(row=row_index, column=5, value=requirement.soglia_minima).font = BODY_FONT
        for col_offset, cell in enumerate(requirement.cells, start=6):
            sheet.cell(
                row=row_index,
                column=col_offset,
                value=f"{cell.esito_label} | {cell.motivazione}",
            ).font = BODY_FONT

    summary_row = 4 + len(matrix.requirements)
    sheet.cell(row=summary_row + 1, column=1, value="Riepilogo").font = Font(bold=True)
    sheet.cell(
        row=summary_row + 1,
        column=2,
        value=(
            f"Soddisfatti: {matrix.summary.get('soddisfatto', 0)} · "
            f"Parziali: {matrix.summary.get('parzialmente', 0)} · "
            f"Non soddisfatti: {matrix.summary.get('non_soddisfatto', 0)}"
        ),
    ).font = BODY_FONT
    _autosize_columns(sheet, max_width=60)
    return _save_workbook(workbook)


def export_participation_xlsx(ctx: TenderExportContext) -> bytes:
    analysis = ctx.participation
    workbook = Workbook()

    summary_sheet = workbook.active
    summary_sheet.title = "Sintesi"
    summary_sheet["A1"] = f"Report partecipabilità – CIG {ctx.tender.cig}"
    summary_sheet["A1"].font = TITLE_FONT
    copertura = analysis.copertura
    summary_rows = [
        ("Forma partecipazione", analysis.forma_label),
        ("Copertura %", f"{copertura.percentuale:.1f}%"),
        ("Requisiti totali", copertura.totale),
        ("Soddisfatti", copertura.soddisfatti),
        ("Parziali", copertura.parziali),
        ("Non soddisfatti", copertura.non_soddisfatti),
        ("Generato il", ctx.exported_at),
    ]
    for row_index, (label, value) in enumerate(summary_rows, start=3):
        summary_sheet.cell(row=row_index, column=1, value=label).font = BODY_FONT
        summary_sheet.cell(row=row_index, column=2, value=value).font = BODY_FONT

    detail_sheet = workbook.create_sheet("Requisiti")
    headers = [
        "ID",
        "Descrizione",
        "Categoria",
        "Esito",
        "Azienda",
        "Posseduto",
        "Richiesto",
        "Motivazione",
        "Critico",
    ]
    for col_index, header in enumerate(headers, start=1):
        detail_sheet.cell(row=1, column=col_index, value=header)
    _style_header_row(detail_sheet, 1)

    for row_index, req in enumerate(analysis.requisiti, start=2):
        detail_sheet.cell(row=row_index, column=1, value=req.requirement_id).font = BODY_FONT
        detail_sheet.cell(row=row_index, column=2, value=req.descrizione).font = BODY_FONT
        detail_sheet.cell(row=row_index, column=3, value=req.categoria).font = BODY_FONT
        detail_sheet.cell(row=row_index, column=4, value=req.esito_label).font = BODY_FONT
        detail_sheet.cell(row=row_index, column=5, value=req.company_name or "").font = BODY_FONT
        detail_sheet.cell(row=row_index, column=6, value=req.valore_posseduto).font = BODY_FONT
        detail_sheet.cell(row=row_index, column=7, value=req.valore_richiesto).font = BODY_FONT
        detail_sheet.cell(row=row_index, column=8, value=req.motivazione).font = BODY_FONT
        detail_sheet.cell(row=row_index, column=9, value="Sì" if req.critico else "No").font = BODY_FONT

    _autosize_columns(summary_sheet)
    _autosize_columns(detail_sheet, max_width=60)
    return _save_workbook(workbook)


def export_relation_xlsx(ctx: TenderExportContext) -> bytes:
    relation = ctx.relation
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Relazione tecnica"
    sheet["A1"] = f"Relazione tecnica – CIG {ctx.tender.cig}"
    sheet["A1"].font = TITLE_FONT
    sheet["A2"] = f"Azienda: {relation.get('company_name') or '—'}"
    sheet["A2"].font = BODY_FONT

    headers = ["Ordine", "Titolo", "Categoria", "Pagine", "Completata", "Contenuto"]
    for col_index, header in enumerate(headers, start=1):
        sheet.cell(row=4, column=col_index, value=header)
    _style_header_row(sheet, 4)

    sections = sorted(relation.get("sections") or [], key=lambda item: item.get("order", 0))
    for row_index, section in enumerate(sections, start=5):
        sheet.cell(row=row_index, column=1, value=section.get("order", "")).font = BODY_FONT
        sheet.cell(row=row_index, column=2, value=section.get("title", "")).font = BODY_FONT
        sheet.cell(row=row_index, column=3, value=section.get("category", "")).font = BODY_FONT
        sheet.cell(row=row_index, column=4, value=section.get("suggested_pages", "")).font = BODY_FONT
        sheet.cell(
            row=row_index,
            column=5,
            value="Sì" if section.get("completed") else "No",
        ).font = BODY_FONT
        content_cell = sheet.cell(row=row_index, column=6, value=section.get("content", ""))
        content_cell.font = BODY_FONT
        content_cell.alignment = Alignment(wrap_text=True, vertical="top")

    _autosize_columns(sheet, max_width=80)
    return _save_workbook(workbook)


def export_economic_xlsx(ctx: TenderExportContext) -> bytes:
    economic = ctx.economic_relation
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Offerta economica"
    sheet["A1"] = f"Offerta economica – CIG {ctx.tender.cig}"
    sheet["A1"].font = TITLE_FONT
    sheet["A2"] = f"Azienda: {economic.get('company_name') or '—'}"
    sheet["A2"].font = BODY_FONT

    outline = economic.get("outline") or {}
    sheet["A3"] = (
        f"Modello: {outline.get('pricing_model', '—')} · "
        f"Importo base: {outline.get('importo_base', '—')}"
    )
    sheet["A3"].font = BODY_FONT

    headers = [
        "Ordine",
        "Voce",
        "U.M.",
        "Quantità",
        "Prezzo unitario",
        "Importo",
        "Ribasso %",
        "Note",
    ]
    for col_index, header in enumerate(headers, start=1):
        sheet.cell(row=5, column=col_index, value=header)
    _style_header_row(sheet, 5)

    items = sorted(economic.get("line_items") or [], key=lambda item: item.get("order", 0))
    for row_index, item in enumerate(items, start=6):
        sheet.cell(row=row_index, column=1, value=item.get("order", "")).font = BODY_FONT
        sheet.cell(row=row_index, column=2, value=item.get("voce", "")).font = BODY_FONT
        sheet.cell(row=row_index, column=3, value=item.get("unita_misura", "")).font = BODY_FONT
        sheet.cell(row=row_index, column=4, value=item.get("quantita", "")).font = BODY_FONT
        sheet.cell(row=row_index, column=5, value=item.get("prezzo_unitario", "")).font = BODY_FONT
        sheet.cell(row=row_index, column=6, value=item.get("importo", "")).font = BODY_FONT
        sheet.cell(row=row_index, column=7, value=item.get("ribasso_percentuale", "")).font = BODY_FONT
        sheet.cell(row=row_index, column=8, value=item.get("notes", "")).font = BODY_FONT

    totals = outline.get("totals") or {}
    total_row = len(items) + 7
    sheet.cell(row=total_row, column=1, value="Totali").font = TITLE_FONT
    sheet.cell(row=total_row, column=6, value=totals.get("totale", "")).font = TITLE_FONT

    _autosize_columns(sheet, max_width=60)
    return _save_workbook(workbook)


def build_xlsx_export(item: str, ctx: TenderExportContext) -> bytes:
    builders = {
        SCHEDA_GARA: export_scheda_xlsx,
        MATRICE_REQUISITI: export_matrix_xlsx,
        REPORT_PARTECIPABILITA: export_participation_xlsx,
        RELAZIONE_TECNICA: export_relation_xlsx,
        OFFERTA_ECONOMICA: export_economic_xlsx,
    }
    return builders[item](ctx)
