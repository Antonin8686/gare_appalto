import csv
import io
import re
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation

from openpyxl import load_workbook

COLUMN_ALIASES: dict[str, tuple[str, ...]] = {
    "cig": ("cig", "codice cig", "cod cig"),
    "cpv": ("cpv", "codice cpv", "cod cpv", "cod. cpv"),
    "importo": ("importo", "valore", "importo base", "importo stimato", "base d'asta"),
    "scadenza": (
        "scadenza",
        "data scadenza",
        "termine",
        "data termine",
        "scadenza presentazione",
        "data scadenza offerte",
    ),
    "stato": ("stato", "status"),
    "oggetto": ("oggetto", "descrizione", "titolo", "denominazione", "oggetto gara"),
}

STATO_MAP = {
    "bozza": "bozza",
    "aperta": "aperta",
    "chiusa": "chiusa",
    "aggiudicata": "aggiudicata",
    "pubblicata": "aperta",
    "in corso": "aperta",
}


@dataclass
class ParsedTenderRow:
    cig: str
    cpv: str
    importo: Decimal
    scadenza: date
    stato: str
    oggetto: str


def _normalize_header(value: str) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def _map_headers(headers: list[str]) -> dict[str, int]:
    normalized = [_normalize_header(header) for header in headers]
    mapping: dict[str, int] = {}

    for field, aliases in COLUMN_ALIASES.items():
        for index, header in enumerate(normalized):
            if header in aliases:
                mapping[field] = index
                break

    return mapping


def _cell_value(row: list, index: int | None) -> str:
    if index is None or index >= len(row):
        return ""
    value = row[index]
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, date):
        return value.strftime("%d/%m/%Y")
    return str(value).strip()


def _parse_importo(value: str) -> Decimal:
    cleaned = value.strip()
    if not cleaned:
        return Decimal("0")

    cleaned = cleaned.replace("€", "").replace("EUR", "").strip()
    cleaned = cleaned.replace(" ", "")

    if "," in cleaned and "." in cleaned:
        if cleaned.rfind(",") > cleaned.rfind("."):
            cleaned = cleaned.replace(".", "").replace(",", ".")
        else:
            cleaned = cleaned.replace(",", "")
    elif "," in cleaned:
        cleaned = cleaned.replace(",", ".")

    try:
        return Decimal(cleaned)
    except InvalidOperation:
        return Decimal("0")


def _parse_scadenza(value: str) -> date:
    cleaned = value.strip()
    if not cleaned:
        return date.today() + timedelta(days=30)

    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d.%m.%Y"):
        try:
            return datetime.strptime(cleaned, fmt).date()
        except ValueError:
            continue

    return date.today() + timedelta(days=30)


def _parse_stato(value: str) -> str:
    cleaned = value.strip().lower()
    if not cleaned:
        return "aperta"
    return STATO_MAP.get(cleaned, "aperta")


def _parse_row(row: list, mapping: dict[str, int]) -> ParsedTenderRow | None:
    cig = _cell_value(row, mapping.get("cig")).upper()
    if not cig:
        return None

    cpv = _cell_value(row, mapping.get("cpv")) or "00000000"
    importo = _parse_importo(_cell_value(row, mapping.get("importo")))
    scadenza = _parse_scadenza(_cell_value(row, mapping.get("scadenza")))
    stato = _parse_stato(_cell_value(row, mapping.get("stato")))
    oggetto = _cell_value(row, mapping.get("oggetto"))

    return ParsedTenderRow(
        cig=cig[:10],
        cpv=cpv[:8],
        importo=importo,
        scadenza=scadenza,
        stato=stato,
        oggetto=oggetto[:500],
    )


def _rows_from_csv(content: bytes) -> list[list[str]]:
    text = content.decode("utf-8-sig")
    reader = csv.reader(io.StringIO(text))
    return [list(row) for row in reader]


def _rows_from_xlsx(content: bytes) -> list[list]:
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    return [list(row) for row in sheet.iter_rows(values_only=True)]


def parse_import_file(content: bytes, filename: str) -> list[ParsedTenderRow]:
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

    if ext == "csv":
        rows = _rows_from_csv(content)
    elif ext in ("xls", "xlsx"):
        rows = _rows_from_xlsx(content)
    else:
        raise ValueError("Formato file non supportato. Usa CSV, XLS o XLSX.")

    if not rows:
        raise ValueError("Il file è vuoto.")

    headers = [str(cell) if cell is not None else "" for cell in rows[0]]
    mapping = _map_headers(headers)
    if "cig" not in mapping:
        raise ValueError("Colonna CIG non trovata nel file.")

    parsed: list[ParsedTenderRow] = []
    for row in rows[1:]:
        if not any(cell not in (None, "") for cell in row):
            continue
        item = _parse_row(list(row), mapping)
        if item:
            parsed.append(item)

    if not parsed:
        raise ValueError("Nessuna gara valida trovata nel file.")

    return parsed
